#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ais_bench 自动性能测试脚本 (JSON 用例驱动)

用法:
    # performance 模式: 用 JSON 用例文件驱动, 支持多数据集类型/输入长度/输出长度/并发/request_rate/pfx 组合
    python3 run_perf.py --cases cases.json

    # agent 模式: 跑原生 SWE-bench, 支持 lite / verified / full / multilingual
    python3 run_perf.py --mode agent --agent-dataset lite --agent-count 10

    # accuracy 模式: 跑 evalscope 精度测试, 当前默认 GPQA Diamond
    python3 run_perf.py --mode accuracy

    # performance 简易模式: 直接命令行指定 (向后兼容)
    python3 run_perf.py -i 32768 -c 1 8 16 --max-out-len 1024 --request-rate 0.5
    python3 run_perf.py -i 32768 -c 1 8 16 --dataset-type gsm sharegpt  # 多数据集类型
    python3 run_perf.py -i 32768 -c 1 8 16 --skip-run       # 只解析已有输出

JSON 用例文件格式 (序号->用例串 映射):
    {
      "1": "32768-1024-4-0.5",           # 默认 sharegpt
      "2": "gsm:32768-1024-4-0.5",       # gsm 数据集
      "3": "sharegpt:16384-1024-8-0-90", # sharegpt + pfx 前缀缓存
      "4": "swebench:16384-1024-16-1"    # swebench 数据集
    }
    也兼容列表形式: ["32768-1024-4-0.5", "gsm:32768-1024-4-0.5"]
    也兼容对象形式: {"in": 32768, "out": 1024, "concurrency": 4, "rate": 0.5, "dataset_type": "gsm"}

用例串格式:  [dataset_type:]输入长度-输出长度-并发-request_rate[-pfx]
  - dataset_type: 可选, gsm/sharegpt/swebench, 省略默认 sharegpt
  - 输入长度:    数据集名里的 in{LEN}
  - 输出长度:    写入 vllm_api_general_stream.py 的 max_out_len
  - 并发:       即 batch_size; 数据集条数 bs = 并发 * 4
  - request_rate: 写入 vllm_api_general_stream.py 的 request_rate (0=尽量打满)
  - pfx (可选): 前缀缓存重复率%, 指定则用 {数据集}-in{LEN}-bs{BS}-pfx{N}.jsonl
                省略或 0 则用普通数据集 {数据集}-in{LEN}-bs{BS}.jsonl
  序号由 JSON 的 key 提供, 仅用于排序/标识。

不带 --cases 时走简易模式: 直接改本文件顶部的 INPUT_LEN/DATASET_TYPES/CONCURRENCIES 等,
输出长度/请求速率/pfx 用命令行 --max-out-len --request-rate --pfx --dataset-type 覆盖。

数据集 jsonl 在脚本所在文件夹下查找。若找不到, 自动调用 process_dataset.py 生成。
启动前会自动检查 ais_bench benchmark 和本次用到的原始数据集；缺失时自动下载/安装，已存在则跳过。
输出(request_rate/max_out_len/path/model/host_ip/host_port)写入 vllm_api_general_stream.py。
最终所有用例结果按 (数据集类型, 输入长度, pfx) 分组写 excel, 保存在执行时的当前文件夹下。
"""

import argparse
import csv
import glob
import json
import os
import re
import signal
import subprocess
import sys
import time
import shutil
import zipfile


# ============================================================
#  配置文件加载 (run_perf.cfg)
#  所有可修改的参数放在同目录下的 run_perf.cfg 中，不再改脚本本身。
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CFG_PATH = os.path.join(SCRIPT_DIR, "run_perf.cfg")
SETUP_SWEBENCH_SCRIPT = os.path.join(SCRIPT_DIR, "setup_swebench.sh")

_DEFAULT_CFG = {
    "mode": "performance",
    "agent": {
        "dataset": "lite",
        "count": 1,
        "step_limit": 200,
        "work_dir": "outputs/agent",
        "run_mode": "all",
        "model_cfg_params": {},
        "auto_prepare": True,
        "benchmark_repo": "https://gh-proxy.com/https://github.com/AISBench/benchmark.git",
        "benchmark_dir": "",
        "benchmark_ref": "",
        "install_requirements": True,
    },
    "accuracy": {
        "dataset": "gpqa_diamond",
        "eval_batch_size": 8,
        "work_dir": "outputs/accuracy",
        "auto_prepare": True,
        "model_cfg_params": {},
        "generation_config": {
            "temperature": 1.0,
            "top_p": 0.95,
            "max_tokens": 130000,
            "timeout": 900,
            "retries": 2,
        },
        "dataset_args": {
            "gpqa_diamond": {
                "filters": {
                    "remove_until": "</think>"
                }
            }
        },
    },
    "performance": {
        # Performance 公共配置
        "kind": "concat",
        "model_cfg_params": {},
        "auto_prepare": True,
        "benchmark_repo": "https://gh-proxy.com/https://github.com/AISBench/benchmark.git",
        "benchmark_dir": "",
        "benchmark_ref": "",
        "install_requirements": True,
        "raw_dataset_dir": "raw_datasets",
        "gsm8k_url": "http://opencompass.oss-cn-shanghai.aliyuncs.com/datasets/data/gsm8k.zip",
        "sharegpt_url": "https://hf-mirror.com/datasets/anon8231489123/ShareGPT_Vicuna_unfiltered/resolve/main/ShareGPT_V3_unfiltered_cleaned_split.json",
        "swebench_url": "https://hf-mirror.com/datasets/princeton-nlp/SWE-bench/resolve/main/data/test-00000-of-00001.parquet",

        # concat 拼接压测专属配置
        "concat": {
            "dataset_dir": "datasets/performance",
            "result_dir": "results/performance",
            "input_len": [32768],
            "concurrencies": [1, 8, 16],
            "default_max_out_len": None,
            "default_request_rate": None,
            "default_pfx": None,
            "dataset_types": ["sharegpt"],
            "raw_gsm_path": "",
            "raw_sharegpt_path": "",
            "raw_swebench_path": "",
        },

        # native_multiturn 原生多轮专属配置
        "native_multiturn": {
            "conversation_count": 100,
            "infer_mode": "every",
            "concurrencies": [1, 8, 16],
            "max_out_len": 512,
            "request_rate": 0,
            "work_dir": "outputs/performance/native_multiturn",
            "result_dir": "results/performance/native_multiturn",
            "raw_sharegpt_path": "",
            "generation_kwargs": {
                "temperature": 0.01,
                "ignore_eos": False,
            },
        },
    },
}


def _deep_merge(base, override):
    """递归合并嵌套配置。"""
    result = dict(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = _deep_merge(result[key], value)
        else:
            result[key] = value
    return result


def _strip_jsonc_comments(text):
    """去掉 JSONC 里的 // 和 /* */ 注释，但保留字符串里的 //。"""
    out = []
    i = 0
    n = len(text)
    in_string = False
    escaped = False

    while i < n:
        ch = text[i]

        if in_string:
            out.append(ch)
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == '"':
                in_string = False
            i += 1
            continue

        if ch == '"':
            in_string = True
            out.append(ch)
            i += 1
            continue

        if ch == "/" and i + 1 < n and text[i + 1] == "/":
            i += 2
            while i < n and text[i] != "\n":
                i += 1
            out.append(" ")
            continue

        if ch == "/" and i + 1 < n and text[i + 1] == "*":
            i += 2
            while i + 1 < n and not (text[i] == "*" and text[i + 1] == "/"):
                if text[i] == "\n":
                    out.append("\n")
                i += 1
            i += 2
            out.append(" ")
            continue

        out.append(ch)
        i += 1

    return "".join(out)


_CFG_TOP_LEVEL_KEYS = {"mode", "agent", "accuracy", "performance"}
_SUPPORTED_MODES = ("performance", "agent", "accuracy")


def _parse_mode_list(value, source="mode"):
    """解析运行模式，支持字符串、逗号分隔、空白分隔和数组。"""
    if isinstance(value, str):
        raw_values = [value]
    elif isinstance(value, (list, tuple)):
        raw_values = list(value)
    else:
        raise RuntimeError("{} 仅支持字符串或数组，当前类型: {}".format(source, type(value).__name__))

    modes = []
    for item in raw_values:
        if not isinstance(item, str):
            raise RuntimeError("{} 里的模式必须是字符串，当前类型: {}".format(source, type(item).__name__))
        for mode in re.split(r"[,\s]+", item.strip()):
            if not mode:
                continue
            if mode not in _SUPPORTED_MODES:
                raise RuntimeError(
                    "{} 仅支持 {} 的任意组合，当前: {}".format(
                        source, " / ".join(_SUPPORTED_MODES), mode
                    )
                )
            if mode not in modes:
                modes.append(mode)

    if not modes:
        raise RuntimeError("{} 不能为空".format(source))
    return modes


def _load_cfg():
    """从 run_perf.cfg 加载配置，支持 JSONC 注释。"""
    if os.path.exists(CFG_PATH):
        with open(CFG_PATH, "r", encoding="utf-8") as f:
            raw = f.read()
        cfg = json.loads(_strip_jsonc_comments(raw))
    else:
        cfg = {}

    unknown_keys = sorted(set(cfg) - _CFG_TOP_LEVEL_KEYS)
    if unknown_keys:
        raise RuntimeError(
            "run_perf.cfg 最外层只支持 mode / agent / accuracy / performance，发现多余字段: {}".format(
                ", ".join(unknown_keys)
            )
        )

    cfg["mode"] = _parse_mode_list(cfg.get("mode", "performance"), "run_perf.cfg mode")
    for section in ("agent", "accuracy", "performance"):
        if section in cfg and not isinstance(cfg[section], dict):
            raise RuntimeError("run_perf.cfg 里 {} 必须是对象".format(section))

    performance_cfg = _deep_merge(_DEFAULT_CFG["performance"], cfg.get("performance", {}))
    allowed_performance_keys = {
        "kind", "model_cfg_params", "auto_prepare",
        "benchmark_repo", "benchmark_dir", "benchmark_ref", "install_requirements",
        "raw_dataset_dir", "gsm8k_url", "sharegpt_url", "swebench_url",
        "concat", "native_multiturn",
    }
    unknown_performance_keys = sorted(set(performance_cfg) - allowed_performance_keys)
    if unknown_performance_keys:
        raise RuntimeError(
            "performance 下只允许公共字段、concat 和 native_multiturn，发现多余字段: {}".format(
                ", ".join(unknown_performance_keys)
            )
        )

    kind = performance_cfg.get("kind", "concat")
    if kind not in ("concat", "native_multiturn"):
        raise RuntimeError(
            "performance.kind 仅支持 concat / native_multiturn，当前: {}".format(kind)
        )

    concat_cfg = performance_cfg.get("concat", {})
    if not isinstance(concat_cfg, dict):
        raise RuntimeError("performance.concat 必须是对象")
    native_cfg = performance_cfg.get("native_multiturn", {})
    if not isinstance(native_cfg, dict):
        raise RuntimeError("performance.native_multiturn 必须是对象")

    allowed_concat_keys = {
        "dataset_dir", "result_dir", "input_len", "concurrencies",
        "default_max_out_len", "default_request_rate", "default_pfx",
        "dataset_types", "raw_gsm_path", "raw_sharegpt_path", "raw_swebench_path",
    }
    unknown_concat_keys = sorted(set(concat_cfg) - allowed_concat_keys)
    if unknown_concat_keys:
        raise RuntimeError(
            "performance.concat 里发现未知字段: {}".format(", ".join(unknown_concat_keys))
        )

    allowed_native_keys = {
        "conversation_count", "infer_mode", "concurrencies", "max_out_len",
        "request_rate", "work_dir", "result_dir", "raw_sharegpt_path",
        "generation_kwargs",
    }
    unknown_native_keys = sorted(set(native_cfg) - allowed_native_keys)
    if unknown_native_keys:
        raise RuntimeError(
            "performance.native_multiturn 里发现未知字段: {}".format(
                ", ".join(unknown_native_keys)
            )
        )

    for subtype, subtype_cfg in (
        ("concat", concat_cfg),
        ("native_multiturn", native_cfg),
    ):
        concurrencies = subtype_cfg.get("concurrencies", [1, 8, 16])
        if not isinstance(concurrencies, (list, tuple)) or not concurrencies:
            raise RuntimeError(
                "performance.{}.concurrencies 必须是非空数组".format(subtype)
            )
        for concurrency in concurrencies:
            if not isinstance(concurrency, int) or concurrency <= 0:
                raise RuntimeError(
                    "performance.{}.concurrencies 里的并发必须是正整数".format(subtype)
                )

    generation_kwargs = native_cfg.get("generation_kwargs", {})
    if not isinstance(generation_kwargs, dict):
        raise RuntimeError("performance.native_multiturn.generation_kwargs 必须是对象")
    infer_mode = native_cfg.get("infer_mode", "every")
    if infer_mode not in ("every", "last", "every_with_gt"):
        raise RuntimeError(
            "performance.native_multiturn.infer_mode 仅支持 every / last / every_with_gt，当前: {}".format(infer_mode)
        )
    try:
        conversation_count = int(native_cfg.get("conversation_count", 100))
    except (TypeError, ValueError) as exc:
        raise RuntimeError("performance.native_multiturn.conversation_count 必须是整数")
    if conversation_count < 0:
        raise RuntimeError("performance.native_multiturn.conversation_count 不能小于 0")
    for field in ("max_out_len", "request_rate"):
        if native_cfg.get(field) is not None:
            try:
                float(native_cfg[field])
            except (TypeError, ValueError):
                raise RuntimeError(
                    "performance.native_multiturn.{} 必须是数字".format(field)
                )

    return _deep_merge(_DEFAULT_CFG, cfg)


_cfg = _load_cfg()

RUN_MODES = _parse_mode_list(_cfg.get("mode", ["performance"]), "run_perf.cfg mode")
RUN_MODE = RUN_MODES[0]
_AGENT_CFG = _cfg.get("agent", {})
_PERFORMANCE_CFG = _cfg.get("performance", {})
_ACCURACY_CFG = _cfg.get("accuracy", {})

# Agent 模式配置
AGENT_DATASET = _AGENT_CFG.get("dataset", "lite")
AGENT_COUNT = int(_AGENT_CFG.get("count", 1))
AGENT_STEP_LIMIT = int(_AGENT_CFG.get("step_limit", 200))
AGENT_WORK_DIR = _AGENT_CFG.get("work_dir", "outputs/agent")
AGENT_RUN_MODE = _AGENT_CFG.get("run_mode", "all")
AGENT_AUTO_PREPARE = bool(_AGENT_CFG.get("auto_prepare", True))
_AGENT_MODEL_CFG_PARAMS = dict(_AGENT_CFG.get("model_cfg_params", {}))

# Performance 公共配置
PERFORMANCE_AUTO_PREPARE = bool(_PERFORMANCE_CFG.get("auto_prepare", True))
PERFORMANCE_KIND = _PERFORMANCE_CFG.get("kind", "concat")
_PERFORMANCE_MODEL_CFG_PARAMS = dict(_PERFORMANCE_CFG.get("model_cfg_params", {}))

# Performance concat 专属配置
_CONCAT_CFG = _PERFORMANCE_CFG.get("concat", {})
PERFORMANCE_DATASET_DIR = os.path.join(
    SCRIPT_DIR,
    _CONCAT_CFG.get("dataset_dir", "datasets/performance"),
)
PERFORMANCE_RESULT_DIR = os.path.join(
    SCRIPT_DIR,
    _CONCAT_CFG.get("result_dir", "results/performance"),
)
INPUT_LEN = _CONCAT_CFG.get("input_len", [32768])
CONCURRENCIES = _CONCAT_CFG.get("concurrencies", [1, 8, 16])
DEFAULT_MAX_OUT_LEN = _CONCAT_CFG.get("default_max_out_len")
DEFAULT_REQUEST_RATE = _CONCAT_CFG.get("default_request_rate")
DEFAULT_PFX = _CONCAT_CFG.get("default_pfx")
DATASET_TYPES = _CONCAT_CFG.get("dataset_types", ["sharegpt"])

# Performance native_multiturn 专属配置
_NATIVE_MULTITURN_CFG = _PERFORMANCE_CFG.get("native_multiturn", {})
NATIVE_CONVERSATION_COUNT = int(_NATIVE_MULTITURN_CFG.get("conversation_count", 100))
NATIVE_INFER_MODE = _NATIVE_MULTITURN_CFG.get("infer_mode", "every")
NATIVE_CONCURRENCIES = _NATIVE_MULTITURN_CFG.get("concurrencies", [1, 8, 16])
NATIVE_MAX_OUT_LEN = _NATIVE_MULTITURN_CFG.get("max_out_len", 512)
NATIVE_REQUEST_RATE = _NATIVE_MULTITURN_CFG.get("request_rate", 0)
NATIVE_WORK_DIR = _NATIVE_MULTITURN_CFG.get(
    "work_dir", "outputs/performance/native_multiturn"
)
NATIVE_RESULT_DIR = os.path.join(
    SCRIPT_DIR,
    _NATIVE_MULTITURN_CFG.get(
        "result_dir", "results/performance/native_multiturn"
    ),
)
NATIVE_RAW_SHAREGPT_PATH = _NATIVE_MULTITURN_CFG.get("raw_sharegpt_path", "")
NATIVE_GENERATION_KWARGS = _NATIVE_MULTITURN_CFG.get(
    "generation_kwargs", {"temperature": 0.01, "ignore_eos": False}
)

# Accuracy 精度测试配置
ACCURACY_DATASET = _ACCURACY_CFG.get("dataset", "gpqa_diamond")
ACCURACY_EVAL_BATCH_SIZE = int(_ACCURACY_CFG.get("eval_batch_size", 8))
ACCURACY_WORK_DIR = _ACCURACY_CFG.get("work_dir", "outputs/accuracy")
ACCURACY_AUTO_PREPARE = bool(_ACCURACY_CFG.get("auto_prepare", True))
_ACCURACY_MODEL_CFG_PARAMS = dict(_ACCURACY_CFG.get("model_cfg_params", {}))
ACCURACY_GENERATION_CONFIG = _ACCURACY_CFG.get(
    "generation_config",
    {
        "temperature": 1.0,
        "top_p": 0.95,
        "max_tokens": 130000,
        "timeout": 900,
        "retries": 2,
    },
)
ACCURACY_DATASET_ARGS = _ACCURACY_CFG.get(
    "dataset_args",
    {"gpqa_diamond": {"filters": {"remove_until": "</think>"}}},
)


def _model_cfg_for_mode(mode):
    """按模式返回对应的模型配置，避免不同模式互相混用。"""
    return {
        "agent": _AGENT_MODEL_CFG_PARAMS,
        "accuracy": _ACCURACY_MODEL_CFG_PARAMS,
        "performance": _PERFORMANCE_MODEL_CFG_PARAMS,
    }.get(mode, _PERFORMANCE_MODEL_CFG_PARAMS)


def _benchmark_cfg_for_mode(mode):
    """按模式返回 benchmark 配置；accuracy 模式不需要 ais_bench。"""
    if mode == "agent":
        return _AGENT_CFG
    if mode == "performance":
        return _PERFORMANCE_CFG
    return {
        "benchmark_repo": "",
        "benchmark_ref": "",
        "benchmark_dir": "",
        "install_requirements": False,
    }


def _set_active_mode(mode):
    """切换当前运行模式，并加载该模式专属的模型和 benchmark 配置。"""
    global RUN_MODE, MODEL_CFG_PARAMS
    global BENCHMARK_REPO, BENCHMARK_REF, INSTALL_REQUIREMENTS, BENCHMARK_DIR

    if mode not in _SUPPORTED_MODES:
        raise RuntimeError("不支持的运行模式: {}".format(mode))
    RUN_MODE = mode
    MODEL_CFG_PARAMS = dict(_model_cfg_for_mode(mode))

    benchmark_cfg = _benchmark_cfg_for_mode(mode)
    BENCHMARK_REPO = benchmark_cfg.get("benchmark_repo", "")
    BENCHMARK_REF = benchmark_cfg.get("benchmark_ref", "")
    INSTALL_REQUIREMENTS = bool(benchmark_cfg.get("install_requirements", True))
    benchmark_dir_cfg = benchmark_cfg.get("benchmark_dir", "")
    BENCHMARK_DIR = (
        os.path.abspath(benchmark_dir_cfg)
        if os.path.isabs(benchmark_dir_cfg)
        else os.path.join(SCRIPT_DIR, benchmark_dir_cfg)
    ) if benchmark_dir_cfg else os.path.join(SCRIPT_DIR, "benchmark")


# 按当前模式选择模型配置
_set_active_mode(RUN_MODE)

# Performance 公共原始数据下载配置
RAW_DATASET_DIR = os.path.join(
    SCRIPT_DIR,
    _PERFORMANCE_CFG.get("raw_dataset_dir", "raw_datasets"),
)
GSM8K_URL = _PERFORMANCE_CFG.get("gsm8k_url", "")
SHAREGPT_URL = _PERFORMANCE_CFG.get("sharegpt_url", "")
SWEBENCH_URL = _PERFORMANCE_CFG.get("swebench_url", "")


def _resolve_raw_path(value, default_name):
    if value:
        return value if os.path.isabs(value) else os.path.join(SCRIPT_DIR, value)
    return os.path.join(RAW_DATASET_DIR, default_name)


# concat 拼接压测专属原始数据路径
RAW_GSM_PATH = _resolve_raw_path(
    _CONCAT_CFG.get("raw_gsm_path", ""),
    os.path.join("gsm8k", "train.jsonl"),
)
RAW_SHAREGPT_PATH = _resolve_raw_path(
    _CONCAT_CFG.get("raw_sharegpt_path", ""),
    "ShareGPT_V3_unfiltered_cleaned_split.json",
)
RAW_SWEBENCH_PATH = _resolve_raw_path(
    _CONCAT_CFG.get("raw_swebench_path", ""),
    os.path.join("swe-bench", "test-00000-of-00001.parquet"),
)


def _kill_process_group(proc, sig):
    """向子进程所在进程组发送信号。"""
    try:
        os.killpg(os.getpgid(proc.pid), sig)
    except ProcessLookupError:
        pass
    except PermissionError:
        pass


def _wait_with_interrupt(proc):
    """等待子进程，Ctrl+C 时先 SIGINT，再 SIGTERM，最后 SIGKILL。"""
    while True:
        try:
            return proc.wait()
        except KeyboardInterrupt:
            print("\n  [interrupt] 收到 Ctrl+C，正在停止子进程...")
            _kill_process_group(proc, signal.SIGINT)
            try:
                proc.wait(timeout=3)
                return proc.returncode
            except subprocess.TimeoutExpired:
                pass

            print("  [interrupt] 子进程未退出，发送 SIGTERM...")
            _kill_process_group(proc, signal.SIGTERM)
            try:
                proc.wait(timeout=3)
                return proc.returncode
            except subprocess.TimeoutExpired:
                pass

            print("  [interrupt] 子进程仍未退出，发送 SIGKILL...")
            _kill_process_group(proc, signal.SIGKILL)
            proc.wait()
            return proc.returncode


def _run_checked(cmd, cwd=None, label=""):
    """执行外部命令，失败时抛 RuntimeError；Ctrl+C 时强制清理子进程组。"""
    print("  [exec] {}".format(" ".join(map(str, cmd))))
    proc = subprocess.Popen(
        cmd,
        cwd=cwd,
        text=True,
        start_new_session=True,
    )
    returncode = _wait_with_interrupt(proc)
    if returncode != 0:
        raise RuntimeError("{} 退出码 {}: {}".format(label or "命令", returncode, " ".join(map(str, cmd))))
    return proc


def _download_file(url, dest):
    """下载文件到 dest；已存在且非空则跳过。优先 wget 断点续传。"""
    if os.path.exists(dest) and os.path.getsize(dest) > 0:
        print("  [dl ] 已存在: {}".format(dest))
        return dest
    os.makedirs(os.path.dirname(os.path.abspath(dest)), exist_ok=True)
    tmp = dest + ".part"
    print("  [dl ] {} -> {}".format(url, dest))
    if shutil.which("wget"):
        cmd = ["wget", "-c", url, "-O", tmp]
    else:
        cmd = ["curl", "-fL", "--retry", "3", "--retry-delay", "2", url, "-o", tmp]
    proc = subprocess.run(cmd, text=True)
    if proc.returncode != 0 or not os.path.exists(tmp) or os.path.getsize(tmp) == 0:
        raise RuntimeError("下载失败: {} -> {}".format(url, dest))
    os.replace(tmp, dest)
    return dest


def _ensure_raw_file(path, url, label):
    """确保单个原始数据文件存在；不存在则自动下载。"""
    if os.path.exists(path) and os.path.getsize(path) > 0:
        print("  [raw ] {} 已存在: {}".format(label, path))
        return path
    if not PERFORMANCE_AUTO_PREPARE:
        raise RuntimeError("{} 不存在且 performance.auto_prepare=false: {}".format(label, path))
    if not url:
        raise RuntimeError("{} 不存在且未配置下载 URL".format(label))
    _download_file(url, path)
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        raise RuntimeError("{} 下载后仍不可用: {}".format(label, path))
    print("  [raw ] {} 已下载: {}".format(label, path))
    return path


def _ensure_gsm8k(path):
    """确保 GSM8K 原始 jsonl 存在；不存在则下载 zip 并解压。"""
    if os.path.exists(path) and os.path.getsize(path) > 0:
        print("  [raw ] GSM8K 已存在: {}".format(path))
        return path
    if not PERFORMANCE_AUTO_PREPARE:
        raise RuntimeError("GSM8K 不存在且 performance.auto_prepare=false: {}".format(path))
    if not GSM8K_URL:
        raise RuntimeError("GSM8K 不存在且未配置 gsm8k_url")
    os.makedirs(RAW_DATASET_DIR, exist_ok=True)
    zip_path = os.path.join(RAW_DATASET_DIR, "gsm8k.zip")
    _download_file(GSM8K_URL, zip_path)
    print("  [raw ] 解压 GSM8K: {}".format(zip_path))
    with zipfile.ZipFile(zip_path) as zf:
        zf.extractall(RAW_DATASET_DIR)
    extracted = os.path.join(RAW_DATASET_DIR, "gsm8k", "train.jsonl")
    if not os.path.exists(extracted):
        extracted = os.path.join(RAW_DATASET_DIR, "gsm8k", "test.jsonl")
    if not os.path.exists(extracted):
        raise RuntimeError("GSM8K 下载解压后未找到 train.jsonl/test.jsonl")
    # 用户指定了自定义路径时，软链到解压结果，保持配置语义不变
    if path != extracted and not os.path.exists(path):
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        os.symlink(extracted, path)
        print("  [raw ] GSM8K symlink: {} -> {}".format(extracted, path))
        return path
    return extracted


def prepare_raw_datasets(dataset_types):
    """按本次实际用到的数据集类型准备原始文件，已存在则跳过。"""
    global RAW_GSM_PATH, RAW_SHAREGPT_PATH, RAW_SWEBENCH_PATH
    for dt in dataset_types:
        if dt == "gsm":
            RAW_GSM_PATH = _ensure_gsm8k(RAW_GSM_PATH)
        elif dt == "sharegpt":
            RAW_SHAREGPT_PATH = _ensure_raw_file(
                RAW_SHAREGPT_PATH, SHAREGPT_URL, "ShareGPT"
            )
        elif dt == "swebench":
            RAW_SWEBENCH_PATH = _ensure_raw_file(
                RAW_SWEBENCH_PATH, SWEBENCH_URL, "SWE-bench"
            )
        else:
            raise RuntimeError("未知数据集类型: {}".format(dt))

# -------------------- Agent 模式: 原生 SWE-bench --------------------
_AGENT_DATASET_HF_ID = {
    "lite": "princeton-nlp/SWE-Bench_Lite",
    "verified": "princeton-nlp/SWE-Bench_Verified",
    "full": "princeton-nlp/SWE-Bench",
    "multilingual": "SWE-bench/SWE-bench_Multilingual",
}


def _import_or_none(name):
    try:
        return __import__(name)
    except Exception:
        return None


def _ensure_agent_environment():
    '''确保 agent 模式需要的 mini-swe-agent / swebench 已安装。'''
    if AGENT_AUTO_PREPARE and os.path.isfile(SETUP_SWEBENCH_SCRIPT):
        print("  [agent] 执行环境配置脚本: {}".format(SETUP_SWEBENCH_SCRIPT))
        _run_checked(
            ["bash", SETUP_SWEBENCH_SCRIPT],
            cwd=SCRIPT_DIR,
            label="setup_swebench.sh",
        )

    missing = []
    if _import_or_none("minisweagent") is None:
        missing.append("minisweagent")
    if _import_or_none("swebench") is None:
        missing.append("swebench")
    if not missing:
        print("  [agent] mini-swe-agent / swebench 环境已就绪")
        return

    if not AGENT_AUTO_PREPARE:
        raise RuntimeError("agent 模式缺少依赖 {}，且 agent.auto_prepare=false".format(missing))

    print("  [agent] 缺少依赖 {}，开始自动安装".format(missing))
    if "minisweagent" in missing:
        repo_dir = os.path.join(SCRIPT_DIR, "third_party", "mini-swe-agent")
        if not os.path.isdir(repo_dir):
            os.makedirs(os.path.dirname(repo_dir), exist_ok=True)
            _run_checked(
                ["git", "clone", "--depth", "1",
                 "https://gh-proxy.com/https://github.com/AISBench/mini-swe-agent.git", repo_dir],
                label="git clone mini-swe-agent",
            )
        _run_checked(
            [sys.executable, "-m", "pip", "install", "-e", repo_dir],
            cwd=repo_dir,
            label="pip install mini-swe-agent",
        )
    if "swebench" in missing:
        repo_dir = os.path.join(SCRIPT_DIR, "third_party", "SWE-bench")
        if not os.path.isdir(repo_dir):
            os.makedirs(os.path.dirname(repo_dir), exist_ok=True)
            _run_checked(
                ["git", "clone", "--branch", "v4.1.0", "--depth", "1",
                 "https://gh-proxy.com/https://github.com/SWE-bench/SWE-bench.git", repo_dir],
                label="git clone SWE-bench",
            )
        _run_checked(
            [sys.executable, "-m", "pip", "install", "-e", repo_dir],
            cwd=repo_dir,
            label="pip install SWE-bench",
        )


def _load_agent_instance_ids(dataset_name, count):
    '''按数据集原始顺序取前 count 条 instance_id；count=0 表示全部。'''
    if dataset_name not in _AGENT_DATASET_HF_ID:
        raise ValueError(
            "agent_dataset 仅支持 {}，当前: {}".format(
                sorted(_AGENT_DATASET_HF_ID), dataset_name
            )
        )
    from datasets import load_dataset

    hf_id = _AGENT_DATASET_HF_ID[dataset_name]
    print("  [agent] 加载数据集: {}".format(hf_id))
    ds = load_dataset(hf_id, split="test")
    ids = list(ds["instance_id"])
    total = len(ids)
    if count is None or count <= 0:
        selected = ids
    else:
        selected = ids[:count]
    print("  [agent] 数据集总数={}，本次选择={}，取前 {} 条".format(total, len(selected), len(selected)))
    return selected


def _agent_filter_regex(instance_ids):
    '''把 instance_id 列表转成精确匹配 regex。'''
    import re as _re
    if not instance_ids:
        raise RuntimeError("agent 模式没有选中任何 instance")
    return "^(?:" + "|".join(_re.escape(x) for x in instance_ids) + ")$"


def _build_agent_config(dataset_name, instance_ids, step_limit):
    '''生成原生 SWE-bench 的 ais_bench 配置文件。'''
    from string import Template

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    cfg_dir = os.path.join(SCRIPT_DIR, "outputs", "agent_configs")
    os.makedirs(cfg_dir, exist_ok=True)
    cfg_path = os.path.join(
        cfg_dir,
        "swebench_agent_{}_{}_{}.py".format(dataset_name, len(instance_ids), timestamp),
    )

    model_name = MODEL_CFG_PARAMS.get("model", "")
    host_ip = MODEL_CFG_PARAMS.get("host_ip", "")
    host_port = MODEL_CFG_PARAMS.get("host_port", "")
    if not model_name or not host_ip or not host_port:
        raise RuntimeError(
            "agent 模式需要在 model_cfg_params 里配置 model / host_ip / host_port"
        )
    api_url = "http://{}:{}/v1".format(host_ip, host_port)
    api_key = MODEL_CFG_PARAMS.get("api_key", "dummy") or "dummy"
    filter_regex = _agent_filter_regex(instance_ids)

    template = Template('''from ais_bench.benchmark.datasets import SWEBenchDataset
from ais_bench.benchmark.partitioners import NaivePartitioner
from ais_bench.benchmark.runners import LocalRunner
from ais_bench.benchmark.tasks import SWEBenchInferTask, SWEBenchEvalTask
from ais_bench.benchmark.summarizers import SWEBenchSummarizer

STEP_LIMIT = $step_limit

models = [
    dict(
        attr="local",
        abbr="swebench",
        type="LiteLLMChat",
        model="$model_name",
        api_key="$api_key",
        url="$api_url",
        batch_size=1,
        max_out_len=16384,
        generation_kwargs=dict(
            max_tokens=32768,
            temperature=0.2,
            timeout=600,
        ),
    )
]

datasets = [
    dict(
        type=SWEBenchDataset,
        abbr="swebench_$dataset_name",
        path="",
        name="$dataset_name",
        split="test",
        filter_spec="$filter_regex",
        shuffle=False,
        step_limit=STEP_LIMIT,
    )
]

summarizer = dict(
    attr="accuracy",
    type=SWEBenchSummarizer,
)

infer = dict(
    partitioner=dict(type=NaivePartitioner),
    runner=dict(
        type=LocalRunner,
        task=dict(type=SWEBenchInferTask),
    ),
)

eval = dict(
    partitioner=dict(type=NaivePartitioner),
    runner=dict(
        type=LocalRunner,
        task=dict(type=SWEBenchEvalTask),
    ),
)
''')
    content = template.substitute(
        step_limit=step_limit,
        model_name=model_name,
        api_key=api_key,
        api_url=api_url,
        dataset_name=dataset_name,
        filter_regex=filter_regex,
    )
    with open(cfg_path, "w", encoding="utf-8") as f:
        f.write(content)
    print("  [agent] 配置文件: {}".format(cfg_path))
    return cfg_path


def run_agent_mode():
    '''执行原生 SWE-bench agent 评测。'''
    print("========== Agent 模式: 原生 SWE-bench ==========")
    print("数据集 = {}".format(AGENT_DATASET))
    print("数量 = {}".format("全部" if AGENT_COUNT <= 0 else AGENT_COUNT))
    print("step_limit = {}".format(AGENT_STEP_LIMIT))
    print("运行阶段 = {}".format(AGENT_RUN_MODE))
    print("工作目录 = {}".format(AGENT_WORK_DIR))

    _ensure_agent_environment()
    instance_ids = _load_agent_instance_ids(AGENT_DATASET, AGENT_COUNT)
    if not instance_ids:
        raise RuntimeError("agent 模式没有选中任何 instance")
    print("  [agent] 选中 instance:")
    for iid in instance_ids:
        print("    - {}".format(iid))

    cfg_path = _build_agent_config(AGENT_DATASET, instance_ids, AGENT_STEP_LIMIT)
    work_dir = os.path.join(SCRIPT_DIR, AGENT_WORK_DIR)
    os.makedirs(work_dir, exist_ok=True)

    cmd = [
        "ais_bench",
        cfg_path,
        "--mode", AGENT_RUN_MODE,
        "--work-dir", work_dir,
        "--max-num-workers", "1",
    ]
    print("  [agent] CMD: {}".format(" ".join(cmd)))
    proc = subprocess.Popen(
        cmd,
        cwd=SCRIPT_DIR,
        text=True,
        start_new_session=True,
    )
    returncode = _wait_with_interrupt(proc)
    if returncode != 0:
        raise RuntimeError("ais_bench agent 模式退出码 {}".format(returncode))
    print("  [agent] 原生 SWE-bench 完成")


# -------------------- Accuracy 模式: evalscope 精度测试 --------------------
def _ensure_accuracy_environment():
    """确保 evalscope 可用；缺失时按配置自动安装。"""
    if shutil.which("evalscope"):
        print("  [accuracy] evalscope 环境已就绪")
        return

    if not ACCURACY_AUTO_PREPARE:
        raise RuntimeError("evalscope 不存在，且 accuracy.auto_prepare=false")

    print("  [accuracy] 未找到 evalscope，开始自动安装")
    _run_checked(
        [sys.executable, "-m", "pip", "install", "evalscope"],
        cwd=SCRIPT_DIR,
        label="pip install evalscope",
    )
    if not shutil.which("evalscope"):
        raise RuntimeError(
            "evalscope 安装后仍不可用。请检查 Python 环境和 PATH，或手动执行: pip install evalscope"
        )
    print("  [accuracy] evalscope 安装完成")


def _accuracy_json_arg(value, name):
    """把 dict 转成 CLI JSON 字符串；字符串则校验后原样传给 evalscope。"""
    if isinstance(value, str):
        try:
            json.loads(value)
        except Exception as exc:
            raise RuntimeError("accuracy.{} 不是合法 JSON: {}".format(name, exc))
        return value
    try:
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    except Exception as exc:
        raise RuntimeError("accuracy.{} 序列化失败: {}".format(name, exc))


def run_accuracy_mode():
    """执行 evalscope 精度测试，当前默认 GPQA Diamond。"""
    print("========== Accuracy 模式: evalscope 精度测试 ==========")
    print("数据集 = {}".format(ACCURACY_DATASET))
    print("eval_batch_size = {}".format(ACCURACY_EVAL_BATCH_SIZE))
    print("工作目录 = {}".format(ACCURACY_WORK_DIR))

    _ensure_accuracy_environment()

    model_name = MODEL_CFG_PARAMS.get("model", "")
    host_ip = MODEL_CFG_PARAMS.get("host_ip", "")
    host_port = MODEL_CFG_PARAMS.get("host_port", "")
    api_key = MODEL_CFG_PARAMS.get("api_key", "EMPTY") or "EMPTY"
    if not model_name or not host_ip or not host_port:
        raise RuntimeError(
            "accuracy 模式需要在 accuracy.model_cfg_params 里配置 model / host_ip / host_port"
        )

    api_url = "http://{}:{}/v1".format(host_ip, host_port)
    work_dir = (
        ACCURACY_WORK_DIR
        if os.path.isabs(ACCURACY_WORK_DIR)
        else os.path.join(SCRIPT_DIR, ACCURACY_WORK_DIR)
    )
    os.makedirs(work_dir, exist_ok=True)

    cmd = [
        "evalscope",
        "eval",
        "--model", model_name,
        "--api-url", api_url,
        "--api-key", api_key,
        "--eval-type", "openai_api",
        "--datasets", ACCURACY_DATASET,
        "--eval-batch-size", str(ACCURACY_EVAL_BATCH_SIZE),
        "--generation-config", _accuracy_json_arg(
            ACCURACY_GENERATION_CONFIG, "generation_config"
        ),
        "--dataset-args", _accuracy_json_arg(
            ACCURACY_DATASET_ARGS, "dataset_args"
        ),
    ]
    print("  [accuracy] CMD: {}".format(" ".join(cmd)))
    _run_checked(cmd, cwd=work_dir, label="evalscope accuracy")
    print("  [accuracy] 精度测试完成")


# 路径常量
def _find_existing_ais_bench_root():
    """定位 ais_bench benchmark 安装根目录。

    支持 pip install 和 pip install -e 两种安装方式。
    通过 import 包定位，最可靠，不依赖 pip 输出格式。
    """
    import importlib as _il

    for pkg_name in ("ais_bench_benchmark", "ais_bench"):
        try:
            mod = _il.import_module(pkg_name)
            # mod.__file__ 如 /path/to/ais_bench_benchmark/__init__.py
            # pip install -e 时包目录的父目录就是源码根
            pkg_dir = os.path.dirname(getattr(mod, "__file__", ""))
            # 候选：pkg_dir 本身、pkg_dir 的父目录
            for root in (pkg_dir, os.path.dirname(pkg_dir)):
                if os.path.isdir(os.path.join(root, "benchmark", "configs", "models")):
                    return root
        except Exception:
            pass

    # 回退：pip show
    import subprocess as _sp
    _pip = [sys.executable, "-m", "pip"]
    for pkg in ("ais_bench_benchmark", "ais_bench", "ais-bench-benchmark"):
        try:
            r = _sp.run(_pip + ["show", pkg], capture_output=True, text=True, timeout=10)
            if r.returncode != 0:
                continue
            loc = None
            for line in r.stdout.splitlines():
                line = line.strip()
                if line.startswith("Editable project location:"):
                    loc = line.split(":", 1)[1].strip()
                elif line.startswith("Location:"):
                    loc = line.split(":", 1)[1].strip()
            if loc is None:
                continue
            for root in (loc, os.path.join(loc, pkg)):
                if os.path.isdir(os.path.join(root, "benchmark", "configs", "models")):
                    return root
        except Exception:
            pass

    # 回退：向上查找
    d = SCRIPT_DIR
    for _ in range(6):
        candidate = os.path.join(d, "benchmark", "configs", "models")
        if os.path.isdir(candidate):
            return d
        parent = os.path.dirname(d)
        if parent == d:
            break
        d = parent

    raise RuntimeError(
        "无法定位 AIS_BENCH_ROOT。请确保 ais_bench_benchmark 已安装（pip install -e ./），"
        "或脚本在 ais_bench 目录树下运行。"
    )

def _install_benchmark_repo(repo_dir):
    """安装 ais_bench benchmark 源码仓库。"""
    if not os.path.isfile(os.path.join(repo_dir, "setup.py")) and not os.path.isfile(os.path.join(repo_dir, "pyproject.toml")):
        raise RuntimeError("benchmark 目录存在但不是可用源码仓库: {}".format(repo_dir))
    print("  [bench] 安装 ais_bench: {}".format(repo_dir))
    _run_checked(
        [sys.executable, "-m", "pip", "install", "-e", repo_dir, "--use-pep517"],
        cwd=repo_dir,
        label="pip install ais_bench",
    )
    if INSTALL_REQUIREMENTS:
        for req_name in ("requirements/api.txt", "requirements/extra.txt"):
            req_path = os.path.join(repo_dir, req_name)
            if os.path.isfile(req_path):
                print("  [bench] 安装依赖: {}".format(req_path))
                _run_checked(
                    [sys.executable, "-m", "pip", "install", "-r", req_path],
                    cwd=repo_dir,
                    label="pip install {}".format(req_name),
                )


def _clone_benchmark_repo(repo_dir):
    """克隆 ais_bench benchmark 源码仓库。"""
    if os.path.isfile(os.path.join(repo_dir, "setup.py")) or os.path.isfile(os.path.join(repo_dir, "pyproject.toml")):
        print("  [bench] benchmark 源码已存在: {}".format(repo_dir))
        return
    if os.path.exists(repo_dir) and os.listdir(repo_dir):
        raise RuntimeError("benchmark_dir 非空但不是 benchmark 源码: {}".format(repo_dir))
    os.makedirs(os.path.dirname(os.path.abspath(repo_dir)), exist_ok=True)
    if not BENCHMARK_REPO:
        raise RuntimeError("ais_bench 未安装且未配置 benchmark_repo")
    cmd = ["git", "clone"]
    if BENCHMARK_REF:
        cmd += ["--branch", BENCHMARK_REF, "--depth", "1"]
    else:
        cmd += ["--depth", "1"]
    cmd += [BENCHMARK_REPO, repo_dir]
    print("  [bench] 克隆 ais_bench: {}".format(BENCHMARK_REPO))
    _run_checked(cmd, label="git clone benchmark")


def _ensure_ais_bench_root():
    """定位或自动安装 ais_bench benchmark。"""
    try:
        return _find_existing_ais_bench_root()
    except Exception as exc:
        if not PERFORMANCE_AUTO_PREPARE:
            raise
        print("  [bench] 未找到已安装的 ais_bench，开始自动准备: {}".format(exc))
        _clone_benchmark_repo(BENCHMARK_DIR)
        _install_benchmark_repo(BENCHMARK_DIR)
        try:
            return _find_existing_ais_bench_root()
        except Exception:
            if os.path.isdir(os.path.join(BENCHMARK_DIR, "benchmark", "configs", "models")):
                return BENCHMARK_DIR
            raise


# performance 模式延迟初始化，避免 accuracy 模式强制依赖 ais_bench。
AIS_BENCH_ROOT = None
PROCESS_DATASET_SCRIPT = os.path.join(SCRIPT_DIR, "process_dataset.py")
MODEL_CFG = None
DATASET_DIR = PERFORMANCE_DATASET_DIR          # performance 模式生成的 jsonl 数据集目录
RUN_CWD = SCRIPT_DIR                           # ais_bench 运行目录，不随执行位置变化
OUTPUTS_ROOT = os.path.join(RUN_CWD, "outputs", "default")
EXCEL_PATH = os.path.join(PERFORMANCE_RESULT_DIR, "性能测试结果.xlsx")


def _init_performance_runtime():
    """初始化 performance 模式运行时依赖的 ais_bench 路径。"""
    global AIS_BENCH_ROOT, MODEL_CFG
    AIS_BENCH_ROOT = _ensure_ais_bench_root()
    MODEL_CFG = os.path.join(
        AIS_BENCH_ROOT,
        "benchmark", "configs", "models", "vllm_api", "vllm_api_general_stream.py",
    )
    os.makedirs(PERFORMANCE_RESULT_DIR, exist_ok=True)

# 数据集类型 → (显示名, 文件前缀, 原始路径配置key, process_dataset --datasettype 值)
_DATASET_INFO = {
    "gsm":      ("GSM8K",    "GSM8K-in",    "RAW_GSM_PATH",      "GSM"),
    "sharegpt": ("ShareGPT", "ShareGPT-in", "RAW_SHAREGPT_PATH", "SHAREGPT"),
    "swebench": ("Swebench", "Swebench-in", "RAW_SWEBENCH_PATH", "SWEBENCH"),
}

# ais_bench 运行命令
AIS_BENCH_CMD = [
    "ais_bench",
    "--models", "vllm_api_general_stream",
    "--datasets", "gsm8k_gen_0_shot_cot_str_perf",
    "--mode", "perf",
    "--debug",
]


# -------------------- 工具函数 --------------------
def parse_number(s):
    """从 '25282.1 ms' / '40.5053 token/s' / '0.0396 req/s' / '4' 里取数值."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return s
    m = re.search(r"-?\d+\.?\d*", str(s))
    return float(m.group()) if m else None


# Excel 单元格不允许控制字符 (含 ANSI 转义码 \x1b), 写入前必须剥离
_ILLEGAL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def clean_cell(value):
    """把任意值清洗成可安全写入 Excel 的值 (剥离 ANSI/控制字符)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    return _ILLEGAL_CHARS_RE.sub("", str(value))


def _set_cfg_field(content, field, value):
    """在 vllm_api_general_stream.py 文本里把某字段替换为 value.

    字符串值加引号, 数字/布尔原样写。匹配 `field=...` 形式 (容空格)。
    返回 (new_content, 替换次数)。找不到该字段时替换次数为 0。
    """
    if isinstance(value, str):
        replacement = '{}="{}"'.format(field, value)
    elif isinstance(value, bool):
        replacement = '{}={}'.format(field, str(value))
    else:
        replacement = '{}={}'.format(field, value)
    pattern = re.compile(r'{}\s*=\s*[^,\n]+'.format(re.escape(field)))
    new_content, n = pattern.subn(replacement, content)
    return new_content, n


def set_model_cfg(params):
    """批量改 vllm_api_general_stream.py 里的字段.

    支持字段: batch_size, max_out_len, request_rate, path, model, host_ip, host_port 等。
    params: dict, 值为 None 的字段跳过。一次读、一次写, 避免重复 IO。
    """
    if not params:
        return
    with open(MODEL_CFG, "r", encoding="utf-8") as f:
        content = f.read()
    changed = []
    for field, value in params.items():
        if value is None:
            continue
        content, n = _set_cfg_field(content, field, value)
        if n == 0:
            print("  [cfg] !! 警告: 在 {} 里找不到 {}=".format(MODEL_CFG, field))
        else:
            changed.append("{}={}".format(field, repr(value)))
    with open(MODEL_CFG, "w", encoding="utf-8") as f:
        f.write(content)
    if changed:
        print("  [cfg] 配置 -> {}".format(", ".join(changed)))


def _dataset_filename(dataset_type, input_len, bs, pfx=None):
    """根据数据集类型构造文件名。

    gsm:      GSM8K-in{LEN}-bs{BS}.jsonl  (无 pfx)
    sharegpt: ShareGPT-in{LEN}-bs{BS}.jsonl 或 ShareGPT-in{LEN}-bs{BS}-pfx{N}.jsonl
    swebench: Swebench-in{LEN}-bs{BS}.jsonl 或 Swebench-in{LEN}-bs{BS}-pfx{N}.jsonl
    """
    _, prefix, _, _ = _DATASET_INFO[dataset_type]
    if pfx:
        return "{}{}-bs{}-pfx{}.jsonl".format(prefix, input_len, bs, pfx)
    else:
        return "{}{}-bs{}.jsonl".format(prefix, input_len, bs)


def find_dataset(dataset_type, input_len, concurrency, pfx=None):
    """在 DATASET_DIR 下查找数据集文件, 返回 (文件名, 绝对路径) 或 (None, None).

    bs = 并发 * 4。支持 gsm/sharegpt/swebench 三种数据集类型。
    """
    bs = concurrency * 4
    ds_name = _dataset_filename(dataset_type, input_len, bs, pfx)
    ds_path = os.path.join(DATASET_DIR, ds_name)
    if os.path.exists(ds_path):
        return ds_name, ds_path
    return None, None


def generate_dataset(dataset_type, input_len, concurrency, pfx=None):
    """调用 process_dataset.py 生成数据集, 返回 (文件名, 绝对路径)。

    生成后再次查找确认文件存在; 若仍不存在则抛异常。
    """
    bs = concurrency * 4
    ds_name = _dataset_filename(dataset_type, input_len, bs, pfx)

    _, _, raw_path_key, ds_type_arg = _DATASET_INFO[dataset_type]
    raw_path = globals().get(raw_path_key, "")
    # tokenizer 模型路径：取 model_cfg_params.path，即 cfg 里的 model_cfg_params.path
    model_path = MODEL_CFG_PARAMS.get("path", "")
    if not model_path:
        raise RuntimeError(
            "model_cfg_params.path 未配置, 无法自动生成数据集 (需要 tokenizer)。"
            "请在 run_perf.cfg 中设置 model_cfg_params.path。"
        )

    cmd = [
        "python3", PROCESS_DATASET_SCRIPT,
        "--bs", str(bs),
        "--inputlen", str(input_len),
        "--datasettype", ds_type_arg,
        "--modelpath", model_path,
    ]

    # 数据集类型特有参数
    if dataset_type == "gsm":
        # GSM 数据路径硬编码在 process_dataset.py 中为 ./GSM8K.jsonl
        if raw_path:
            gsm_link = os.path.join(DATASET_DIR, "GSM8K.jsonl")
            if not os.path.exists(gsm_link) and os.path.exists(raw_path):
                os.symlink(raw_path, gsm_link)
                print("  [gen] GSM 原始数据 symlink: {} -> {}".format(raw_path, gsm_link))
    elif dataset_type == "sharegpt":
        if not raw_path:
            raise RuntimeError("RAW_SHAREGPT_PATH 未配置, 无法生成 ShareGPT 数据集")
        cmd += ["--sharegptpath", raw_path]
    elif dataset_type == "swebench":
        if not raw_path:
            raise RuntimeError("RAW_SWEBENCH_PATH 未配置, 无法生成 SWE-bench 数据集")
        cmd += ["--swebenchpath", raw_path]

    # pfx 模式
    if pfx:
        cmd += ["--mode", "prefix", "--prefix_ratio", str(pfx / 100.0)]

    os.makedirs(DATASET_DIR, exist_ok=True)
    print("  [gen] 自动生成数据集: {}".format(ds_name))
    print("  [gen] CMD: {}".format(" ".join(str(x) for x in cmd)))
    proc = subprocess.run(
        cmd,
        cwd=DATASET_DIR,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
    )
    if proc.returncode != 0:
        tail = "\n".join(proc.stdout.splitlines()[-30:])
        raise RuntimeError(
            "process_dataset.py 退出码 {}: \n{}".format(proc.returncode, tail)
        )
    # 生成后验证文件存在
    ds_name, ds_path = find_dataset(dataset_type, input_len, concurrency, pfx)
    if ds_name is None:
        raise RuntimeError("数据集生成后仍未找到: {}".format(ds_name))
    print("  [gen] 数据集已生成: {}".format(ds_path))
    return ds_name, ds_path


def bind_dataset(ds_name):
    """ln -sf {ds_name} -> test.jsonl / train.jsonl (在脚本目录和 ais_bench 期望路径下)。

    ais_bench 内部会去 AIS_BENCH_ROOT/datasets/gsm8k/ 下找 test.jsonl 和 train.jsonl，
    这里同时软链到脚本目录和该路径，避免 FileNotFoundError。
    """
    ds_path = os.path.join(DATASET_DIR, ds_name)
    # 脚本目录下的软链
    for link_name in ("test.jsonl", "train.jsonl"):
        link_path = os.path.join(DATASET_DIR, link_name)
        if os.path.islink(link_path) or os.path.exists(link_path):
            os.remove(link_path)
        os.symlink(ds_name, link_path)
        print("  [ln ] {} -> {}".format(ds_name, link_name))

    # ais_bench 期望路径：AIS_BENCH_ROOT/datasets/gsm8k/test.jsonl
    gsm8k_dir = os.path.join(AIS_BENCH_ROOT, "datasets", "gsm8k")
    os.makedirs(gsm8k_dir, exist_ok=True)
    for link_name in ("test.jsonl", "train.jsonl"):
        link_path = os.path.join(gsm8k_dir, link_name)
        if os.path.islink(link_path) or os.path.exists(link_path):
            os.remove(link_path)
        os.symlink(ds_path, link_path)
        print("  [ln ] {} -> {}".format(ds_path, link_path))


def run_ais_bench():
    """跑一次 ais_bench, 返回 (returncode, stdout+stderr).

    cwd 用 RUN_CWD (当前进程 cwd): ais_bench 输出目录是相对路径 `outputs/default/...`,
    落在哪取决于运行时工作目录, 必须和执行 run_perf.py 的位置一致。
    输出同时打到终端，方便实时观察进度。
    """
    print("  [run] {}  (cwd={})".format(" ".join(AIS_BENCH_CMD), RUN_CWD))
    proc = subprocess.run(
        AIS_BENCH_CMD,
        cwd=RUN_CWD,
        stdout=None,
        stderr=None,
        text=True,
    )
    return proc.returncode, proc.stdout


def find_latest_result(before_dirs, output_root=OUTPUTS_ROOT):
    """在 output_root 下找本次运行新生成的 timestamp 目录, 返回结果目录.

    只看 *本次新增* 的目录: 找不到 performance csv/json 的就返回 None,
    绝不 fallback 到历史目录——否则失败的用例会复用旧结果, 数值张冠李戴.
    """
    if not os.path.isdir(output_root):
        return None
    cur_dirs = set(
        d for d in os.listdir(output_root)
        if os.path.isdir(os.path.join(output_root, d))
    )
    new_dirs = sorted(cur_dirs - before_dirs)
    for ts in reversed(new_dirs):
        performances_root = os.path.join(output_root, ts, "performances")
        if not os.path.isdir(performances_root):
            continue
        for model_dir_name in sorted(os.listdir(performances_root)):
            model_dir = os.path.join(performances_root, model_dir_name)
            if not os.path.isdir(model_dir):
                continue
            for csv_path in glob.glob(os.path.join(model_dir, "*.csv")):
                json_path = os.path.splitext(csv_path)[0] + ".json"
                if os.path.exists(json_path):
                    return model_dir
    return None


def parse_result(result_dir):
    """解析 performance csv + json, 返回 (perf_rows, common_rows, summary_dict)."""
    csv_paths = glob.glob(os.path.join(result_dir, "*.csv"))
    if not csv_paths:
        raise RuntimeError("结果目录里没有 csv: {}".format(result_dir))
    csv_path = csv_paths[0]
    json_path = os.path.splitext(csv_path)[0] + ".json"
    if not os.path.exists(json_path):
        raise RuntimeError("结果目录里没有对应 json: {}".format(json_path))

    perf_rows = []  # [(param, stage, avg, min, max, median, p75, p90, p99, n), ...]
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        for row in reader:
            if not row or not row[0].strip():
                continue
            perf_rows.append(row)

    with open(json_path, "r", encoding="utf-8") as f:
        common = json.load(f)
    common_rows = [(k, list(v.keys())[0], list(v.values())[0]) for k, v in common.items()]

    perf_map = {r[0]: r for r in perf_rows}
    def avg(param):
        return parse_number(perf_map[param][2]) if param in perf_map else None

    summary = {
        "平均E2EL(ms)": avg("E2EL"),
        "平均TTFT(ms)": avg("TTFT"),
        "平均TPOT(ms)": avg("TPOT"),
        "平均ITL(ms)": avg("ITL"),
        "平均输出token吞吐(token/s)": avg("OutputTokenThroughput"),
        "总吞吐(token/s)": parse_number(common.get("Total Token Throughput", {}).get("total")),
        "请求吞吐(req/s)": parse_number(common.get("Request Throughput", {}).get("total")),
        "总请求数": common.get("Total Requests", {}).get("total"),
    }
    return perf_rows, common_rows, summary


# -------------------- 用例解析 --------------------
def parse_case_str(s, seq=""):
    """解析用例串 '[dataset_type:]输入-输出-并发-rate[-pfx]'。

    格式: 可选的 dataset_type: 前缀 (gsm/sharegpt/swebench), 省略默认 sharegpt。
    seq: 用例序号 (来自 JSON 的 key 或简易模式计数), 仅用于标识, 不参与解析。
    返回 dict: {seq, dataset_type, input_len, out_len, concurrency, request_rate, pfx}。
    解析失败抛 ValueError。
    """
    s = str(s).strip()
    # 检查是否带 dataset_type: 前缀
    dataset_type = "sharegpt"  # 默认
    for dt in _DATASET_INFO:
        if s.lower().startswith(dt + ":"):
            dataset_type = dt
            s = s[len(dt) + 1:]
            break
    parts = s.split("-")
    if len(parts) < 4:
        raise ValueError("用例串字段不足 (需至少 输入-输出-并发-rate): {}".format(s))
    try:
        case = {
            "seq": str(seq),
            "dataset_type": dataset_type,
            "input_len": int(parts[0]),
            "out_len": int(parts[1]),
            "concurrency": int(parts[2]),
            "request_rate": float(parts[3]),
            "pfx": None,
        }
    except ValueError as e:
        raise ValueError("用例串数值解析失败 {}: {}".format(s, e))
    if len(parts) >= 5:
        try:
            pfx_val = int(parts[4])
            case["pfx"] = pfx_val if pfx_val else None
        except ValueError:
            raise ValueError("pfx 字段不是整数: {}".format(s))
    return case


def load_cases(path):
    """从 JSON 文件加载用例列表。

    支持两种格式:
      1) 序号->用例串 映射 (推荐):
         {"1": "32768-1024-4-0.5",  "2": "32768-1024-4-0.5-90", ...}
      2) 用例串列表 (兼容):
         ["32768-1024-4-0.5", "32768-1024-4-0.5-90", ...]
    用例串格式: 输入长度-输出长度-并发-request_rate[-pfx]  (序号由 JSON key 提供)
    """
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    cases = []
    if isinstance(data, dict):
        # {"1": "...", "2": "..."}  按 key 排序保持顺序
        for seq in sorted(data.keys(), key=lambda x: (len(str(x)), str(x))):
            item = data[seq]
            if isinstance(item, str):
                cases.append(parse_case_str(item, seq=seq))
            elif isinstance(item, dict):
                cases.append(_case_from_obj(item, seq=seq))
            else:
                raise ValueError("用例值必须是字符串或对象: {}: {}".format(seq, item))
    elif isinstance(data, list):
        for i, item in enumerate(data, start=1):
            if isinstance(item, str):
                cases.append(parse_case_str(item, seq=i))
            elif isinstance(item, dict):
                cases.append(_case_from_obj(item, seq=i))
            else:
                raise ValueError("用例项必须是字符串或对象: {}".format(item))
    else:
        raise ValueError("JSON 用例文件必须是 dict 或 list: {}".format(path))
    return cases


def _case_from_obj(item, seq=""):
    """从对象 {in, out, concurrency, rate, pfx, dataset_type} 构造用例 dict."""
    return {
        "seq": str(item.get("seq", item.get("id", seq))),
        "dataset_type": str(item.get("dataset_type", item.get("ds", "sharegpt"))),
        "input_len": int(item.get("in", item.get("input_len"))),
        "out_len": int(item.get("out", item.get("out_len"))),
        "concurrency": int(item.get("concurrency", item.get("c", 0))),
        "request_rate": float(item.get("rate", item.get("request_rate", 0))),
        "pfx": int(item["pfx"]) if "pfx" in item and item["pfx"] is not None else None,
    }


# -------------------- Excel 写入 --------------------
def safe_sheet_name(name):
    """excel sheet 名最长 31 字符, 且不能含 []:*?/\\."""
    for ch in "[]:*?/\\":
        name = name.replace(ch, "_")
    return name[:31]


def write_excel(cases, out_path=None):
    """cases: list of dict(case_name, batch_size, summary, perf_rows, common_rows, status, error, ...).
    out_path: 输出路径, 默认用全局 EXCEL_PATH。
    """
    out_path = out_path or EXCEL_PATH

    # performance 模式才需要 openpyxl；accuracy 模式不强制依赖 Excel 库。
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()

    # ---- 汇总 sheet ----
    ws = wb.active
    ws.title = "性能测试汇总"
    headers = [
        "配置名称", "数据集类型", "输入长度", "输出长度", "并发数", "request_rate", "pfx",
        "平均E2EL(ms)", "平均TTFT(ms)", "平均TPOT(ms)", "平均ITL(ms)",
        "平均输出token吞吐(token/s)", "总吞吐(token/s)", "请求吞吐(req/s)", "总请求数",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for case in cases:
        s = case["summary"] or {}
        ws.append([
            clean_cell(case["case_name"]),
            clean_cell(case.get("dataset_type", "sharegpt")),
            case.get("input_len"),
            case.get("out_len"),
            case.get("concurrency"),
            case.get("request_rate"),
            case.get("pfx"),
            clean_cell(s.get("平均E2EL(ms)")),
            clean_cell(s.get("平均TTFT(ms)")),
            clean_cell(s.get("平均TPOT(ms)")),
            clean_cell(s.get("平均ITL(ms)")),
            clean_cell(s.get("平均输出token吞吐(token/s)")),
            clean_cell(s.get("总吞吐(token/s)")),
            clean_cell(s.get("请求吞吐(req/s)")),
            clean_cell(s.get("总请求数")),
        ])
    for col in "ABCDEFGHIJK":
        ws.column_dimensions[col].width = 22
    for col in "LMNO":
        ws.column_dimensions[col].width = 22

    # ---- 每个用例一个明细 sheet ----
    for case in cases:
        name = safe_sheet_name(case["case_name"])
        ws = wb.create_sheet(name)
        bold = Font(bold=True)

        if case["status"] != "ok":
            ws["A1"] = clean_cell("用例失败: {}".format(case.get("error", case["status"])))
            ws["A1"].font = bold
            continue

        # 性能参数
        ws["A1"] = "性能参数"
        ws["A1"].font = bold
        perf_header = [
            "Performance Parameters", "Stage", "Average", "Min", "Max",
            "Median", "P75", "P90", "P99", "N",
        ]
        ws.append(perf_header)
        for c in ws[ws.max_row]:
            c.font = bold
        for row in case["perf_rows"]:
            out = [clean_cell(row[0]), clean_cell(row[1])]
            for cell in row[2:]:
                out.append(parse_number(cell))
            ws.append(out)

        ws.append([])
        ws.append(["通用指标"])
        ws[ws.max_row][0].font = bold
        ws.append(["Common Metric", "Stage", "Value"])
        for c in ws[ws.max_row]:
            c.font = bold
        for k, stage, v in case["common_rows"]:
            ws.append([clean_cell(k), clean_cell(stage), parse_number(v) if v is not None else v])

        for col in "ABCDEFGHIJ":
            ws.column_dimensions[col].width = 22

    wb.save(out_path)
    print("\n[done] Excel 已保存: {}".format(out_path))


# -------------------- 主流程 --------------------
def run_case(case, skip_run=False):
    """跑一个用例。case: dict from parse_case_str / load_cases."""
    dataset_type = case.get("dataset_type", "sharegpt")
    input_len = case["input_len"]
    concurrency = case["concurrency"]
    out_len = case["out_len"]
    request_rate = case["request_rate"]
    pfx = case["pfx"]

    bs = concurrency * 4
    # 用例命名: 数据集类型 + 文件名 + 并发
    ds_stub = _dataset_filename(dataset_type, input_len, bs, pfx)
    # 去掉 .jsonl 后缀
    if ds_stub.endswith(".jsonl"):
        ds_stub = ds_stub[:-6]
    case_name = "{}-{}并发".format(ds_stub, concurrency)
    print("\n========== 用例 #{}: {} ==========".format(case.get("seq", ""), case_name))

    result = {
        "case_name": case_name,
        "dataset_type": dataset_type,
        "input_len": input_len,
        "out_len": out_len,
        "concurrency": concurrency,
        "request_rate": request_rate,
        "pfx": pfx,
        "summary": None,
        "perf_rows": [],
        "common_rows": [],
        "status": "failed",
        "error": "",
    }

    try:
        # 1) 查数据集; 找不到则自动生成
        ds_name, ds_path = find_dataset(dataset_type, input_len, concurrency, pfx)
        if ds_name is None:
            print("  [ds ] 数据集不存在, 尝试自动生成...")
            ds_name, ds_path = generate_dataset(dataset_type, input_len, concurrency, pfx)
        print("  [ds ] {}".format(ds_name))

        # 2) 改动态配置 (max_out_len + request_rate + batch_size，每个用例不同)
        cfg = {"max_out_len": out_len, "request_rate": request_rate, "batch_size": concurrency}
        set_model_cfg(cfg)

        # 3) 绑定数据集
        bind_dataset(ds_name)

        before_dirs = set(
            d for d in os.listdir(OUTPUTS_ROOT)
            if os.path.isdir(os.path.join(OUTPUTS_ROOT, d))
        ) if os.path.isdir(OUTPUTS_ROOT) else set()

        if skip_run:
            # 调试: 用最新的已有输出
            result_dir = find_latest_result(set())
        else:
            rc, _ = run_ais_bench()
            if rc != 0:
                raise RuntimeError("ais_bench 退出码 {}".format(rc))
            result_dir = find_latest_result(before_dirs)

        if not result_dir:
            raise RuntimeError("找不到本次输出的 performance 结果目录; 可能 ais_bench 未成功生成")

        perf_rows, common_rows, summary = parse_result(result_dir)
        result["perf_rows"] = perf_rows
        result["common_rows"] = common_rows
        result["summary"] = summary
        result["status"] = "ok"
        print("  [ok ] 结果: {}".format(result_dir))
    except Exception as e:
        result["error"] = str(e)
        result["status"] = "failed"
        print("  [ERR] {}".format(e))

    return result


def _resolve_native_sharegpt_path():
    """解析原生多轮 ShareGPT 数据路径。

    native_multiturn 的数据路径完全独立；留空时使用公共 raw_dataset_dir 下的默认路径，
    不回退 concat 模式的 performance.raw_sharegpt_path。
    """
    if NATIVE_RAW_SHAREGPT_PATH:
        value = NATIVE_RAW_SHAREGPT_PATH
        return value if os.path.isabs(value) else os.path.join(SCRIPT_DIR, value)
    return os.path.join(RAW_DATASET_DIR, "ShareGPT_V3_unfiltered_cleaned_split.json")


def _prepare_native_sharegpt_dataset():
    """准备 AISBench 原生 ShareGPT 数据；count>0 时截取前 N 组有效多轮对话。"""
    raw_path = _ensure_raw_file(
        _resolve_native_sharegpt_path(), SHAREGPT_URL, "ShareGPT"
    )
    if NATIVE_CONVERSATION_COUNT == 0:
        print("  [native] 使用全量 ShareGPT: {}".format(raw_path))
        return raw_path
    if NATIVE_CONVERSATION_COUNT < 0:
        raise RuntimeError("conversation_count 不能小于 0")

    subset_dir = os.path.join(RAW_DATASET_DIR, "sharegpt")
    os.makedirs(subset_dir, exist_ok=True)
    subset_path = os.path.join(
        subset_dir, "native_multiturn_top{}.json".format(NATIVE_CONVERSATION_COUNT)
    )
    print("  [native] 读取 ShareGPT 并截取前 {} 组有效多轮对话: {}".format(
        NATIVE_CONVERSATION_COUNT, raw_path
    ))
    with open(raw_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    valid_conversations = []
    for item in data:
        conversations = item.get("conversations", [])
        if len(conversations) < 2:
            continue
        if len(conversations) % 2 != 0:
            continue
        if conversations[0].get("from") != "human":
            continue
        valid_conversations.append(item)
        if len(valid_conversations) >= NATIVE_CONVERSATION_COUNT:
            break

    if not valid_conversations:
        raise RuntimeError("ShareGPT 里没有找到有效多轮对话: {}".format(raw_path))
    if len(valid_conversations) < NATIVE_CONVERSATION_COUNT:
        print("  [native] 有效多轮对话不足，实际取 {} 组".format(len(valid_conversations)))

    tmp_path = subset_path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(valid_conversations, f, ensure_ascii=False)
    os.replace(tmp_path, subset_path)
    print("  [native] ShareGPT 子集: {}".format(subset_path))
    return subset_path


def _build_native_multiturn_config(concurrency, out_len, request_rate, dataset_path, work_dir):
    """生成 AISBench 原生 ShareGPT 多轮性能测试配置。"""
    model_path = MODEL_CFG_PARAMS.get("path", "")
    model_name = MODEL_CFG_PARAMS.get("model", "")
    host_ip = MODEL_CFG_PARAMS.get("host_ip", "")
    host_port = MODEL_CFG_PARAMS.get("host_port", "")
    if not model_path or not model_name or not host_ip or not host_port:
        raise RuntimeError(
            "native_multiturn 需要在 performance.model_cfg_params 里配置 path / model / host_ip / host_port"
        )

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    cfg_dir = os.path.join(SCRIPT_DIR, "outputs", "performance_configs")
    os.makedirs(cfg_dir, exist_ok=True)
    cfg_path = os.path.join(
        cfg_dir,
        "native_sharegpt_multiturn_bs{}_{}.py".format(concurrency, timestamp),
    )

    content = """from ais_bench.benchmark.models import VLLMCustomAPIChat
from ais_bench.benchmark.datasets import ShareGPTDataset, ShareGPTEvaluator
from ais_bench.benchmark.openicl.icl_prompt_template import MultiTurnPromptTemplate
from ais_bench.benchmark.openicl.icl_retriever import ZeroRetriever
from ais_bench.benchmark.openicl.icl_inferencer import MultiTurnGenInferencer
from ais_bench.benchmark.summarizers.default_perf import DefaultPerfSummarizer
from ais_bench.benchmark.calculators.default_perf_metric_calculator import DefaultPerfMetricCalculator
from ais_bench.benchmark.utils.postprocess.model_postprocessors import extract_non_reasoning_content

models = [
    dict(
        attr="service",
        type=VLLMCustomAPIChat,
        abbr="vllm-multiturn-api-chat-stream",
        path={model_path!r},
        model={model_name!r},
        stream=True,
        request_rate={request_rate!r},
        retry=2,
        api_key={api_key!r},
        host_ip={host_ip!r},
        host_port={host_port!r},
        url="",
        max_out_len={out_len!r},
        batch_size={concurrency!r},
        trust_remote_code=False,
        generation_kwargs={generation_kwargs!r},
        pred_postprocessor=dict(type=extract_non_reasoning_content),
    )
]

datasets = [
    dict(
        abbr="sharegpt",
        type=ShareGPTDataset,
        disable_shuffle=True,
        path={dataset_path!r},
        reader_cfg=dict(
            input_columns=["question", "answer"],
            output_column="answer",
        ),
        infer_cfg=dict(
            prompt_template=dict(
                type=MultiTurnPromptTemplate,
                template=dict(
                    round=[
                        dict(role="HUMAN", prompt="{{question}}"),
                        dict(role="BOT", prompt="{{answer}}"),
                    ]
                ),
            ),
            retriever=dict(type=ZeroRetriever),
            inferencer=dict(
                type=MultiTurnGenInferencer,
                infer_mode={infer_mode!r},
            ),
        ),
        eval_cfg=dict(evaluator=dict(type=ShareGPTEvaluator)),
    )
]

summarizer = dict(
    attr="performance",
    type=DefaultPerfSummarizer,
    calculator=dict(
        type=DefaultPerfMetricCalculator,
        stats_list=["Average", "Min", "Max", "Median", "P75", "P90", "P99"],
    ),
)

work_dir = {work_dir!r}
"""
    with open(cfg_path, "w", encoding="utf-8") as f:
        f.write(content.format(
            model_path=model_path,
            model_name=model_name,
            host_ip=host_ip,
            host_port=host_port,
            api_key=MODEL_CFG_PARAMS.get("api_key", "") or "",
            out_len=out_len,
            concurrency=concurrency,
            request_rate=request_rate,
            generation_kwargs=NATIVE_GENERATION_KWARGS,
            dataset_path=dataset_path,
            infer_mode=NATIVE_INFER_MODE,
            work_dir=work_dir,
        ))
    print("  [native] AISBench 配置: {}".format(cfg_path))
    return cfg_path


def run_native_multiturn_mode(args):
    """执行 AISBench 原生 ShareGPT 多轮对话性能测试。"""
    if args.cases:
        raise RuntimeError("performance.kind=native_multiturn 不支持 --cases，请使用简易模式参数")

    concurrencies = args.concurrency if args.concurrency else NATIVE_CONCURRENCIES
    if args.skip_run and len(concurrencies) != 1:
        raise RuntimeError("native_multiturn 的 --skip-run 目前只支持单个并发")

    _init_performance_runtime()
    dataset_path = _prepare_native_sharegpt_dataset()
    work_dir = (
        NATIVE_WORK_DIR
        if os.path.isabs(NATIVE_WORK_DIR)
        else os.path.join(SCRIPT_DIR, NATIVE_WORK_DIR)
    )
    os.makedirs(work_dir, exist_ok=True)
    os.makedirs(NATIVE_RESULT_DIR, exist_ok=True)

    out_len = args.max_out_len if args.max_out_len is not None else NATIVE_MAX_OUT_LEN
    request_rate = (
        args.request_rate if args.request_rate is not None else NATIVE_REQUEST_RATE
    )
    out_len = int(out_len)
    request_rate = float(request_rate)

    print("数据集类型 = ShareGPT 原生多轮")
    print("conversation_count = {}".format(
        "全部" if NATIVE_CONVERSATION_COUNT == 0 else NATIVE_CONVERSATION_COUNT
    ))
    print("infer_mode = {}".format(NATIVE_INFER_MODE))
    print("并发列表 = {}".format(concurrencies))
    print("输出目录 = {}".format(work_dir))
    print("Excel 目录 = {}".format(NATIVE_RESULT_DIR))

    results = []
    for seq, concurrency in enumerate(concurrencies, start=1):
        case_name = "NativeShareGPT-multiturn-{}并发".format(concurrency)
        print("\n========== 用例 #{}: {} ==========".format(seq, case_name))
        result = {
            "case_name": case_name,
            "dataset_type": "native_sharegpt",
            "input_len": "native",
            "out_len": out_len,
            "concurrency": concurrency,
            "request_rate": request_rate,
            "pfx": None,
            "summary": None,
            "perf_rows": [],
            "common_rows": [],
            "status": "failed",
            "error": "",
        }
        try:
            before_dirs = set(
                d for d in os.listdir(work_dir)
                if os.path.isdir(os.path.join(work_dir, d))
            ) if os.path.isdir(work_dir) else set()

            if args.skip_run:
                result_dir = find_latest_result(set(), output_root=work_dir)
            else:
                cfg_path = _build_native_multiturn_config(
                    concurrency, out_len, request_rate, dataset_path, work_dir
                )
                cmd = [
                    "ais_bench",
                    cfg_path,
                    "--mode", "perf",
                    "--debug",
                    "--work-dir", work_dir,
                ]
                print("  [native] CMD: {}".format(" ".join(cmd)))
                _run_checked(cmd, cwd=SCRIPT_DIR, label="ais_bench native multiturn")
                result_dir = find_latest_result(before_dirs, output_root=work_dir)

            if not result_dir:
                raise RuntimeError("找不到本次输出的 performance 结果目录")
            perf_rows, common_rows, summary = parse_result(result_dir)
            result.update({
                "perf_rows": perf_rows,
                "common_rows": common_rows,
                "summary": summary,
                "status": "ok",
            })
            print("  [ok ] 结果: {}".format(result_dir))
        except Exception as exc:
            result["error"] = str(exc)
            result["status"] = "failed"
            print("  [ERR] {}".format(exc))
        results.append(result)

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    excel_path = args.excel or os.path.join(
        NATIVE_RESULT_DIR,
        "性能测试结果_native_sharegpt_multiturn_{}.xlsx".format(timestamp),
    )
    write_excel(results, out_path=excel_path)

    print("\n========== Native ShareGPT 多轮汇总 ==========")
    print("{:<40} {:>8} {:>8} {:>10} {:>10} {:>8}".format(
        "用例", "并发", "rate", "E2EL", "总吞吐", "状态"))
    for result in results:
        summary = result["summary"] or {}
        print("{:<40} {:>8} {:>8} {:>10} {:>10} {:>8}".format(
            result["case_name"][:40],
            str(result["concurrency"]),
            str(result["request_rate"]),
            str(summary.get("平均E2EL(ms)", "-")),
            str(summary.get("总吞吐(token/s)", "-")),
            result["status"],
        ))


def run_performance_mode(args):
    """执行 performance 模式：concat 拼接压测或 native_multiturn 原生多轮压测。"""
    global EXCEL_PATH

    if PERFORMANCE_KIND == "native_multiturn":
        run_native_multiturn_mode(args)
        return

    _init_performance_runtime()
    os.makedirs(PERFORMANCE_RESULT_DIR, exist_ok=True)
    if args.excel:
        EXCEL_PATH = args.excel

    default_out_len = args.max_out_len if args.max_out_len is not None else DEFAULT_MAX_OUT_LEN
    default_request_rate = (
        args.request_rate if args.request_rate is not None else DEFAULT_REQUEST_RATE
    )

    print("数据集目录 = {}".format(DATASET_DIR))
    print("执行目录(cwd) = {}".format(RUN_CWD))
    print("Excel 输出 = {}".format(EXCEL_PATH))
    active = {k: v for k, v in MODEL_CFG_PARAMS.items() if v is not None}
    print("模型配置(生效) = {}".format(active if active else "(全用文件现有值)"))
    # 静态配置（path/model/host_ip/host_port）只在启动时写一次，整个测试过程不变
    set_model_cfg(MODEL_CFG_PARAMS)

    # ---- 构造用例列表 ----
    if args.cases:
        cases = load_cases(args.cases)
        print("模式: JSON 用例驱动, 共 {} 条用例".format(len(cases)))
    else:
        # 简易模式: 输入长度 × 数据集类型 × 并发 × pfx 四重笛卡尔积
        input_lens = args.input_len if args.input_len else INPUT_LEN
        dataset_types = args.dataset_type if args.dataset_type else DATASET_TYPES
        concurrencies = args.concurrency if args.concurrency else CONCURRENCIES
        pfxs = args.pfx if args.pfx else [DEFAULT_PFX]
        # pfx: 0 视为普通数据集 (None)
        pfxs = [p if p else None for p in pfxs]
        print("模式: 简易 (数据集类型 {} × 输入长度 {} × 并发 {} × pfx {})".format(
            dataset_types, input_lens, concurrencies, [p if p is not None else "无" for p in pfxs]))
        cases = []
        seq = 1
        for dt in dataset_types:
            for il in input_lens:
                for c in concurrencies:
                    for pfx in pfxs:
                        cases.append({
                            "seq": str(seq),
                            "dataset_type": dt,
                            "input_len": il,
                            "out_len": default_out_len,
                            "concurrency": c,
                            "request_rate": default_request_rate,
                            "pfx": pfx,
                        })
                        seq += 1

    # ---- 自动准备本次用到的原始数据集 ----
    active_dataset_types = sorted({case.get("dataset_type", "sharegpt") for case in cases})
    print("performance 自动准备 = {}".format("开启" if PERFORMANCE_AUTO_PREPARE else "关闭"))
    prepare_raw_datasets(active_dataset_types)

    # ---- 跑用例 ----
    results = []
    for case in cases:
        results.append(run_case(case, skip_run=args.skip_run))

    # ---- 按 (数据集类型, 输入长度, pfx) 分组, 每组写一个 Excel ----
    groups = {}
    for r in results:
        key = (r.get("dataset_type", "sharegpt"), r["input_len"], r["pfx"])
        groups.setdefault(key, []).append(r)

    # 时间戳，避免覆盖之前的结果
    timestamp = time.strftime("%Y%m%d_%H%M%S")

    for (dt, input_len, pfx), group_results in groups.items():
        pfx_tag = "_pfx{}".format(pfx) if pfx else ""
        group_path = os.path.join(
            PERFORMANCE_RESULT_DIR,
            "性能测试结果_{}_in{}{}_{}.xlsx".format(dt, input_len, pfx_tag, timestamp),
        )
        write_excel(group_results, out_path=group_path)

    # 终端汇总
    print("\n========== 汇总 ==========")
    print("{:<40} {:>8} {:>8} {:>10} {:>10} {:>8}".format(
        "用例", "并发", "rate", "E2EL", "总吞吐", "状态"))
    for r in results:
        summary = r["summary"] or {}
        print("{:<40} {:>8} {:>8} {:>10} {:>10} {:>8}".format(
            r["case_name"][:40],
            str(r["concurrency"]),
            str(r["request_rate"]),
            str(summary.get("平均E2EL(ms)", "-")),
            str(summary.get("总吞吐(token/s)", "-")),
            r["status"],
        ))


def _apply_mode_overrides(args, mode):
    """把通用命令行模型参数覆盖到当前模式自己的模型配置上。"""
    global MODEL_CFG_PARAMS

    MODEL_CFG_PARAMS = dict(MODEL_CFG_PARAMS)
    if args.model:
        MODEL_CFG_PARAMS["model"] = args.model
    if args.host_ip:
        MODEL_CFG_PARAMS["host_ip"] = args.host_ip
    if args.host_port is not None:
        MODEL_CFG_PARAMS["host_port"] = args.host_port
    if args.api_key and (mode in ("agent", "accuracy") or (mode == "performance" and PERFORMANCE_KIND == "native_multiturn")):
        MODEL_CFG_PARAMS["api_key"] = args.api_key
    if args.path and mode == "performance":
        MODEL_CFG_PARAMS["path"] = args.path


def main():
    ap = argparse.ArgumentParser(description="xllm 自动测试入口：性能 / Agent / 精度")
    ap.add_argument("--cases", help="performance 模式: JSON 用例文件路径")
    ap.add_argument(
        "--mode",
        nargs="+",
        action="extend",
        default=None,
        help=(
            "运行模式，支持 performance / agent / accuracy 任意组合；"
            "可写 --mode agent performance，也可写 --mode agent,performance"
        ),
    )
    ap.add_argument("--agent-dataset", choices=sorted(_AGENT_DATASET_HF_ID), default=None,
                    help="agent 模式: SWE-bench 数据集类型")
    ap.add_argument("--agent-count", type=int, default=None,
                    help="agent 模式: 取前 N 条; 0 表示全部")
    ap.add_argument("--agent-run-mode", choices=["infer", "eval", "all"], default=None,
                    help="agent 模式: infer=只推理, eval=只评测, all=推理+评测")
    ap.add_argument("--accuracy-dataset", default=None,
                    help="accuracy 模式: evalscope 数据集，默认 accuracy.dataset")
    ap.add_argument("--accuracy-batch-size", type=int, default=None,
                    help="accuracy 模式: eval-batch-size，默认 accuracy.eval_batch_size")
    ap.add_argument("--accuracy-work-dir", default=None,
                    help="accuracy 模式: 输出目录，默认 accuracy.work_dir")
    ap.add_argument("--performance-kind", choices=["concat", "native_multiturn"], default=None,
                    help="performance 模式: concat=拼接压测, native_multiturn=AISBench 原生 ShareGPT 多轮")
    ap.add_argument("--native-conversation-count", type=int, default=None,
                    help="native_multiturn: 取前 N 组有效多轮对话; 0=全部")
    ap.add_argument("--native-infer-mode", choices=["every", "last", "every_with_gt"], default=None,
                    help="native_multiturn: MultiTurnGenInferencer 模式")
    ap.add_argument("--native-work-dir", default=None,
                    help="native_multiturn: AISBench 输出目录")
    ap.add_argument("--native-result-dir", default=None,
                    help="native_multiturn: Excel 输出目录")
    ap.add_argument("--native-raw-sharegpt-path", default=None,
                    help="native_multiturn: 原始 ShareGPT JSON 路径")
    # 简易模式 (向后兼容)
    ap.add_argument("-i", "--input-len", type=int, nargs="+", default=None,
                    help="简易模式: 输入长度列表")
    ap.add_argument("-c", "--concurrency", type=int, nargs="+", default=None,
                    help="performance 通用覆盖: 并发数列表；concat/native 各自使用自己的默认值")
    ap.add_argument("--max-out-len", type=int, default=None,
                    help="performance 通用覆盖: 输出长度；concat/native 各自使用自己的默认值")
    ap.add_argument("--request-rate", type=float, default=None,
                    help="performance 通用覆盖: request_rate；concat/native 各自使用自己的默认值")
    ap.add_argument("--pfx", type=int, nargs="+", default=None,
                    help="concat 模式: 前缀缓存重复率%% 列表；native_multiturn 忽略")
    ap.add_argument("--dataset-type", dest="dataset_type", nargs="+", default=None,
                    choices=["gsm", "sharegpt", "swebench"],
                    help="简易模式: 数据集类型列表 (gsm/sharegpt/swebench), 默认用脚本顶部 DATASET_TYPES")
    # 模型配置覆盖
    ap.add_argument("--path", dest="path", default=None, help="模型权重路径")
    ap.add_argument("--model", dest="model", default=None, help="模型名")
    ap.add_argument("--host-ip", dest="host_ip", default=None, help="服务 IP")
    ap.add_argument("--host-port", dest="host_port", type=int, default=None, help="服务端口")
    ap.add_argument("--api-key", dest="api_key", default=None, help="agent / accuracy / native_multiturn 模式 OpenAI API Key")
    ap.add_argument("--excel", default=None, help="Excel 输出路径 (默认当前文件夹下)")
    ap.add_argument("--skip-run", action="store_true", help="只解析已有输出, 不重新跑")
    args = ap.parse_args()

    global RUN_MODE, RUN_MODES
    global AGENT_DATASET, AGENT_COUNT, AGENT_STEP_LIMIT, AGENT_WORK_DIR
    global ACCURACY_DATASET, ACCURACY_EVAL_BATCH_SIZE, ACCURACY_WORK_DIR
    global PERFORMANCE_KIND, NATIVE_CONVERSATION_COUNT, NATIVE_INFER_MODE
    global NATIVE_WORK_DIR, NATIVE_RESULT_DIR, NATIVE_RAW_SHAREGPT_PATH

    if args.performance_kind:
        PERFORMANCE_KIND = args.performance_kind
    if args.native_conversation_count is not None:
        NATIVE_CONVERSATION_COUNT = args.native_conversation_count
    if args.native_infer_mode:
        NATIVE_INFER_MODE = args.native_infer_mode
    if args.native_work_dir:
        NATIVE_WORK_DIR = args.native_work_dir
    if args.native_result_dir:
        NATIVE_RESULT_DIR = (
            args.native_result_dir
            if os.path.isabs(args.native_result_dir)
            else os.path.join(SCRIPT_DIR, args.native_result_dir)
        )
    if args.native_raw_sharegpt_path:
        NATIVE_RAW_SHAREGPT_PATH = args.native_raw_sharegpt_path
    if PERFORMANCE_KIND not in ("concat", "native_multiturn"):
        raise RuntimeError("performance.kind 仅支持 concat / native_multiturn")
    if NATIVE_CONVERSATION_COUNT < 0:
        raise RuntimeError("native conversation_count 不能小于 0")

    if args.agent_dataset:
        AGENT_DATASET = args.agent_dataset
    if args.agent_count is not None:
        AGENT_COUNT = args.agent_count
    if args.agent_run_mode:
        AGENT_RUN_MODE = args.agent_run_mode
    if args.accuracy_dataset:
        ACCURACY_DATASET = args.accuracy_dataset
    if args.accuracy_batch_size is not None:
        ACCURACY_EVAL_BATCH_SIZE = args.accuracy_batch_size
    if args.accuracy_work_dir:
        ACCURACY_WORK_DIR = args.accuracy_work_dir

    cli_modes = _parse_mode_list(args.mode, "命令行 --mode") if args.mode else []
    run_modes = cli_modes or RUN_MODES
    print("运行模式 = {}（按顺序执行）".format(" -> ".join(run_modes)))

    for mode in run_modes:
        print("\n========== 混合运行切换: {} ==========".format(mode))
        _set_active_mode(mode)
        _apply_mode_overrides(args, mode)
        if mode == "agent":
            run_agent_mode()
        elif mode == "accuracy":
            run_accuracy_mode()
        elif mode == "performance":
            run_performance_mode(args)
        else:
            raise RuntimeError("不支持的运行模式: {}".format(mode))

if __name__ == "__main__":
    main()
